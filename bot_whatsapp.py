import openpyxl
import webbrowser
import pyautogui
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from urllib.parse import quote
from time import sleep

def enviar_mensagem(nome_da_pessoa, telefone, vencimento):
    mensagem = f'texto da mensagem'
    link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={telefone}&text={quote(mensagem)}'
    webbrowser.open(link_mensagem_whatsapp);
    sleep(30)
    seta = pyautogui.locateCenterOnScreen('seta_whatsapp.png')
    sleep(5)
    pyautogui.click(seta[0],seta[1])
    sleep(5)
    pyautogui.hotkey('ctrl','w')
    sleep(3)

def processar_planilha(planilha):
    for linha in planilha.iter_rows(min_row=2):
        nome_da_pessoa = linha[0].value
        telefone = linha[1].value
        vencimento = linha[2].value
        enviar_mensagem(nome_da_pessoa, telefone, vencimento)

def processar_arquivo(nome_planilha):
    try:
        workbook = openpyxl.load_workbook('nome_da_planilha.xlsx')
        pagina_app = workbook[nome_planilha]
        processar_planilha(pagina_app)
        messagebox.showinfo('Sucesso', 'Mensagens enviadas com sucesso')
    except Exception as e:
        messagebox.showerror('Erro', f'Erro ao processar arquivo: {str(e)}')

app = tk.Tk()
app.title('texto do app')

sheet_names = []

def carregar_nomes_planilhas():
    global sheet_names
    workbook = openpyxl.load_workbook('nome_da_planilha.xlsx')
    sheet_names = workbook.sheetnames

carregar_nomes_planilhas()

selected_sheet = tk.StringVar(value=sheet_names[0])

ttk.Label(app, text='texto que precede o botao:').grid(column=0, row=0, padx=10, pady=10)
planilha_combobox = ttk.Combobox(app, textvariable=selected_sheet, values=sheet_names, state="readonly")
planilha_combobox.grid(column=1, row=0)

ttk.Button(app, text='texto do botão', command=lambda: processar_arquivo(selected_sheet.get())).grid(column=2, row=0)

app.mainloop()